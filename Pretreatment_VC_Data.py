"""
数据预处理
1、处理数据数据集中重复的数据，并作删除操作
2、对数据中无用的数据进行处理
"""
import re
import openpyxl

# 处理数据函数 将无用字符删除，并提取出有用的数据信息
def Retreatment(readfile,writefile):
    lines_seen = set()
    outfile = open(writefile, "w")
    f = open(readfile, "r")
    # 无用的数据
    for line in f:
        if line not in lines_seen:
            lines_seen.add(line)
            new_line = re.sub(r'[^A-Za-z0-9]+', ' ', line)
            str_list = new_line.split()
            #print(str_list)
            #['500', 'network', '0', 'router', '5', '5', '13431', '3', '2', 'Input', '3', '1', '2', '1', '0', '0', '0', '0', '0', '0', '0', '0', '0', '1', '0', '0', '1', '0', '0', '0', '0', '1', '2', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '1', '0', '0', '1', '0', '0', '1', '0', '0', '0', '0', '0', '0', '0', '0', '0', '0', '1']
            # 必会删除的数据：'network'下标为1；'0'下标为2；'router'下标为3；'13431'传输的数据下标为6；'Input'下标为9；'3'输入单元编号下标为10（跟前面重复）
            # 可选择删除的数据：'3' 输入单元编号下标为7；'2'输入单元下的VC通道编号下标为8
            ans = [1, 2, 3, 6, 9, 10]
            cnt = 0
            for a in ans:
                del str_list[a - cnt]
                cnt += 1
            # 保存数据
            for i in range(len(str_list)):
                s = str(str_list[i]).replace('[', '').replace(']', '') + " "  # 去除[],这两行按数据不同，可以选择
                outfile.write(s)
            outfile.write('\n')
    outfile.close()

# 将从TXT文件读出的数据保存到EXCEl文件中：
def write_excel(readfile, writefile):
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    # f = xlwt.Workbook()
    # sheet1 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
    read_file = open(readfile, "r")
    i = 0
    for line in read_file:
        new_line = re.sub(r'[^A-Za-z0-9]+', ' ', line)
        str_list = new_line.split()
        # print(str_list)
        # 将数据写入第 i 行，第 j 列
        for j in range(len(str_list)):
            # print(str_list[j])
            outws.write(i, j, str_list[j])
        i = i + 1
    outwb.save(writefile)  # 保存文件




if __name__ == '__main__':
    # 从原始数据中，去除重复数据，以及提取出需要的属性字段
    # 要是想获取对应的属性字段，控制这个变量
    # ans = [1, 2, 3, 6, 7, 8, 9, 10]
    # 这个数据值要删除的没有用的属性字段，切记！！！ 是删除无用字段 不要搞错哦！！！
    # 上面那个数组将路由器的输入编号属性 7下标 和对应输入单元编号的VC编号属性 8下标 删除了，要想保留可以删除上面数组的 7和8
    # 这里唯一不好的是，对文件路径设计的比较多。
    Retreatment('./data/uniform_vc_buffer_occupancy.txt', './process_data/uniform_router.txt')
    #### 决定还是不吧数据转换成EXCEL文件了。由于EXCEL文件对数据的条数是有限制的，不方便后续对数据的处理
    # write_excel('./process_data/uniform_router.txt', './process_data/uniform_router.xls')

    print ("success")



